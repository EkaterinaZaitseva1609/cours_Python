import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


def save_list(newline):
  # book = openpyxl.load_workbook('my_book.xlsx')
  book = load_workbook('my_book.xlsx')
  sheet = book.active

  sheet['A1'] = 'Имя'
  sheet['B1'] = 'Номер телефона'

  row = 2
  sheet[row][0].value = newline[0]
  sheet[row][1].value = newline[1]
  row += 1
  book.save("my_book.xlsx")
  book.close()





